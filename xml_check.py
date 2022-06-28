# These need to be specified before running
# What String IDs are called in the relevant XML file
string_id = "resname"
# What the target language attribute is called in the relevant XML file
target_language = "target-language"
# What the tag is called that the target language attribute sits in
target_lang_where = "file"
# What segments are called in the relevant XML file
segment = "trans-unit"


############# Do not touch below this line #############

import openpyxl
import re
from bs4 import BeautifulSoup as bs

class TransUnit:
    def __init__(self, stringID, source, target, lang):
        self.stringID = stringID
        self.source = source
        self.target = target
        self.lang = lang


def convertFromXML():
    global string_id
    global target_language
    global target_lang_where
    global segment
    garbage_heap = []
    trans_units = []
    def getInput():
        input_name = input("Please enter the file name of the XML file you want to convert.\n")
        try:
            with open(input_name, 'r', encoding="utf-8") as fhand:
                data = fhand.read()
                return data
        except:
            print("That didn't work. Please try again.")
            return getInput()
    data = getInput()
    bs_data = bs(data, 'xml')
    lang = bs_data.find(target_lang_where).get(target_language)
    for tag in bs_data.find_all(segment):
        stringID = tag.get(string_id)
        sourceRaw = str(tag.find('source'))
        if re.search('\<source\>(.*)\<\/source\>', sourceRaw, re.DOTALL):
            source = re.search('\<source\>(.*)\<\/source\>', sourceRaw, re.DOTALL).group(1)
        else:
            garbage_heap.append(sourceRaw)
            continue
        targetRaw = str(tag.find('target'))
        if re.search('\<target\>(.*)\<\/target\>', targetRaw, re.DOTALL):
            target = re.search('\<target\>(.*)\<\/target\>', targetRaw, re.DOTALL).group(1)
        trans_units.append(TransUnit(stringID, source, target, lang))
    print("Completed reading trans units.\nPrinting garbage heap.\nIf you see actual text in here, something went wrong.")
    print(garbage_heap)
    return trans_units


def writeToExcel(file_name = None):
    trans_units = convertFromXML()
    file_name = file_name
    def getSheet():
        nonlocal file_name
        if file_name:
            return openpyxl.load_workbook(file_name)
        prompt = input("Would you like to create a new workbook or populate an existing one?\nEnter 1 for new, or the file name of the existing workbook.\n")
        if prompt == "1":
            wb = openpyxl.Workbook()
            return wb
        else:
            try:
                wb = openpyxl.load_workbook(prompt)
                file_name = prompt
                return wb
            except:
                print("That didn't work. Please try again.")
                return getSheet()
    def continueWorking():
        prompt = input("Would you like to add additional XML data into the same sheet?\n1 for Yes\nAnything else for No\n")
        if prompt == "1":
            return writeToExcel(file_name)
    wb = getSheet()
    sheet = wb.active
    next_row = 1
    column = 1
    if not file_name:
        sheet.cell(column=1, row=next_row, value="String ID")
        sheet.cell(column=2, row=next_row, value="English")
        sheet.cell(column=3, row=next_row, value=trans_units[0].lang)
        next_row += 1
        for unit in trans_units:
            try:
                sheet.cell(column=1, row=next_row, value=unit.stringID)
                sheet.cell(column=2, row=next_row, value=unit.source)
                sheet.cell(column=3, row=next_row, value=unit.target)
                next_row +=1
            except:
                print("Something broke writing to the new sheet.")
        file_name = input("Please enter a name for the output Excel sheet (including the extension).\n")
        wb.save(file_name)
        continueWorking()
    else:
        occupied_columns = []
        for cell in sheet[1]:
            occupied_columns.append(cell.value)
        column += len(occupied_columns)
        sheet.cell(column=column, row=next_row, value=trans_units[0].lang)
        next_row += 1
        for unit in trans_units:
            if unit.stringID != sheet.cell(next_row, 1).value:
                print("Warning: Trans unit's string ID does not match. Row: {}\nChecking if we can find it...".format(next_row))
                for unit2 in trans_units:
                    if unit2.stringID == sheet.cell(next_row, 1).value:
                        print("Found a match.")
                        sheet.cell(column=column, row=next_row, value=unit2.target)
                next_row += 1
            sheet.cell(column=column, row=next_row, value=unit.target)
            next_row += 1
        wb.save(file_name)
        continueWorking()

def convertFromExcel(file_name = None):
    trans_units = []
    file_name = file_name
    def getWB():
        nonlocal file_name
        if file_name:
            return openpyxl.load_workbook(file_name)
        prompt = input("Please enter the Excel file to read from.\n")
        try:
            wb = openpyxl.load_workbook(prompt)
            file_name = prompt
            return wb
        except:
            print("That didn't work. Please try again.")
            return getWB()
    wb = getWB()
    sheet = wb.active
    next_row = 1
    column = 1
    langs = []
    for cell in sheet[1]:
        if cell.value != "String ID" and cell.value != "English":
            langs.append(cell.value)
    def whichLang():
        nonlocal langs
        print("Which language would you like to populate to XML?")
        for lang in langs:
            print("{}: {}".format(lang, langs.index(lang)))
        prompt = (input("Please enter the language number.\n"))
        try:
            selection = langs[int(prompt)]
            return int(prompt)
        except:
            print("That doesn't seem to be a valid selection.\nPlease try again.")
            return whichLang()
    lang_index = whichLang()
    lang = langs[lang_index]
    column += lang_index + 2
    column_header = openpyxl.utils.cell.get_column_letter(column)
    for row, cellObj in enumerate(list(sheet.columns)[column - 1]):
        if row == 0:
            continue
        else:
            source = sheet.cell(column=2, row=row+1).value
            target = sheet.cell(column=column, row=row+1).value
            stringID = sheet.cell(column=1, row=row+1).value
            trans_units.append(TransUnit(stringID, source, target, lang))
    print("Finished reading segments.")
    return [trans_units, file_name]

def writeToXML(excel_file_name = None):
    excel_file_name = excel_file_name
    if not excel_file_name:
        stuff = convertFromExcel()
        excel_file_name = stuff[1]
        trans_units = stuff[0]
    else:
        stuff = convertFromExcel(excel_file_name)
        trans_units = stuff[0]
    xml_file_name = None
    def continueWorking():
        prompt = input("Would you like to read additional languages from this sheet?\n1 for Yes\nAnything else for No\n")
        if prompt == "1":
            return writeToXML(excel_file_name)
    def getInput():
        nonlocal xml_file_name
        input_name = input("Please enter the file name of the XML file you want to write to.\n")
        try:
            with open(input_name, 'r', encoding="utf-8") as fhand:
                data = fhand.read()
                xml_file_name = input_name
                return data
        except:
            print("That didn't work. Please try again.")
            return getInput()
    data = getInput()
    bs_data = bs(data, 'xml')
    if bs_data.find(target_lang_where).get(target_language) != trans_units[0].lang:
        print("WARNING! Target language in XML file does not match translations data.")
        return writeToXML(excel_file_name)
    else:
        for tag in bs_data.find_all(segment):
            for unit in trans_units:
                if unit.stringID == tag.get(string_id):
                    if re.search('\<source\>(.*)\<\/source\>', str(tag.get('source')), re.DOTALL):
                        if re.search('\<source\>(.*)\<\/source\>', str(tag.get('source')), re.DOTALL).strip() != unit.source:
                            print("WARNING! XML source\n{}\ndoes not match trans unit source:\n{}".format(tag.source.string, unit.source))
                    else:
                        try:
                            tag.target.string = unit.target
                            trans_units.remove(unit)
                        except:
                            print("Replacing target broke.")
                            print(unit.stringID)
                            continue
        print("Writing to XML done.\n{} trans units left unused.".format(len(trans_units)))
        with open(xml_file_name, "r", encoding="utf-8") as fhand:
            original_data = fhand.read()
        list1 = str(bs_data).split('>\n')
        list2 = str(original_data).split('>\n')
        if len(list1) != len(list2):
            print("Arrays aren't the same length!")
        else:
            for i in range(len(list1) - 1):
                if not re.search('\<target\>(.*)\<\/target', list1[i], re.DOTALL):
                    continue
                else:
                    target = re.search('\<target\>(.*)\<\/target', list1[i], re.DOTALL).group(1)
                    old_target = re.search('\<target\>(.*)\<\/target', list2[i], re.DOTALL).group(1)
                    list2[i] = re.sub(old_target, target, list2[i], re.DOTALL)
        new_data = '>\n'.join(list2)
        with open("test.xlf", "w", encoding="utf-8") as fhand2:
            fhand2.write(new_data)
        fhand.close()
        fhand2.close()


def wrapperUI():
    while True:
        prompt = input("What would you like to do?\n1 - Convert from XML to multilingual Excel\n2 - Convert from Excel back to XML\n3 - Quit\n")
        if prompt == "1":
            writeToExcel()
        if prompt == "2":
            writeToXML()
        if prompt == "3":
            break

wrapperUI()
